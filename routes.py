from datetime import date
from email.mime import image, message
import os
from flask import render_template, redirect, url_for, flash,make_response,send_file,request, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import login_user, logout_user, login_required, current_user
from app import app
from flask import current_app
from flask_mail import Message
from models import Expense, create_notification, db, User,Income,Budget,Investment, Goal,Notification,Feedback
from forms import ExpenseForm, ExpenseForm, RegisterForm, LoginForm, IncomeForm,BudgetForm,InvestmentForm,GoalForm
from sqlalchemy import func,desc,extract
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from flask_mail import Message
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import *

@app.route("/")
def home():
    return redirect(url_for("login"))

@app.route("/register", methods=["GET", "POST"])
def register():
    form = RegisterForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()
        if user:
            flash("Email already exists!", "danger")
            return redirect(url_for("register"))
        hashed_password = generate_password_hash(
            form.password.data,
            method="pbkdf2:sha256"
        )
        new_user = User(
            fullname=form.fullname.data,
            email=form.email.data,
            password=hashed_password
        )
        db.session.add(new_user)
        db.session.commit()
        flash("Registration Successful!", "success")
        return redirect(url_for("login"))
    return render_template("register.html", form=form)


@app.route("/login", methods=["GET", "POST"])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()
        if user and check_password_hash(user.password, form.password.data):
            login_user(user)
            flash("Login Successful!", "success")
            return redirect(url_for("dashboard"))
        else:
            flash("Invalid Email or Password", "danger")
    return render_template("login.html", form=form)

    
@app.route("/dashboard")
@login_required
def dashboard(): 
    total_income = db.session.query(
        func.sum(Income.amount)
    ).filter(
        Income.user_id == current_user.id
    ).scalar() or 0
    total_expense = db.session.query(
        func.sum(Expense.amount)
    ).filter(
        Expense.user_id == current_user.id
    ).scalar() or 0
    total_budget = db.session.query(
        func.sum(Budget.budget_amount)
    ).filter(
        Budget.user_id == current_user.id
    ).scalar() or 0
    total_invested = db.session.query(
        func.sum(Investment.amount)
    ).filter(
        Investment.user_id == current_user.id
    ).scalar() or 0
    total_investment = db.session.query(
        func.sum(Investment.current_value)
    ).filter(
        Investment.user_id == current_user.id
    ).scalar() or 0
    total_savings = total_income - total_expense  
    income_growth = 12.5
    expense_growth = 4.3
    savings_ratio = round(
        (total_savings / total_income) * 100,
        2
    ) if total_income > 0 else 0
    roi = round(
        (
            (total_investment - total_invested)
            / total_invested
        ) * 100,
        2
    ) if total_invested > 0 else 0   
    health_score = 92
    if health_score >= 90:
        health_status = "Excellent"
    elif health_score >= 75:
        health_status = "Good"
    elif health_score >= 60:
        health_status = "Fair"
    else:
        health_status = "Poor"    
    investment_growth = 12.8
    debt_ratio = 18
    expense_ratio = round(
        (total_expense / total_income) * 100,
        2
    ) if total_income > 0 else 0
    recommendations = [
        "Increase monthly SIP by ₹2,000.",
        "Maintain savings above 30% of income.",
        "Reduce discretionary spending.",
        "Continue maintaining a low debt ratio.",
        "Build an emergency fund covering 6 months of expenses."
    ]
    recent_income = Income.query.filter_by(
        user_id=current_user.id
    ).order_by(
        desc(Income.id)
    ).limit(5).all()
    recent_expense = Expense.query.filter_by(
        user_id=current_user.id
    ).order_by(
        desc(Expense.expense_id)
    ).limit(5).all()
    recent_budget = Budget.query.filter_by(
        user_id=current_user.id
    ).order_by(
        desc(Budget.id)
    ).limit(5).all()
    expense_chart = db.session.query(
        Expense.category,
        func.sum(Expense.amount)
    ).filter(
        Expense.user_id == current_user.id
    ).group_by(
        Expense.category
    ).all()
    labels = []
    values = []
    for category, amount in expense_chart:
        labels.append(category)
        values.append(float(amount))
    months = [
        "Jan","Feb","Mar","Apr","May","Jun",
        "Jul","Aug","Sep","Oct","Nov","Dec"
    ]
    income_data = [0] * 12
    expense_data = [0] * 12
    income_result = db.session.query(
        extract("month", Income.income_date),
        func.sum(Income.amount)
    ).filter(
        Income.user_id == current_user.id
    ).group_by(
        extract("month", Income.income_date)
    ).all()
    for month, amount in income_result:
        income_data[int(month)-1] = float(amount)
    expense_result = db.session.query(
        extract("month", Expense.expense_date),
        func.sum(Expense.amount)
    ).filter(
        Expense.user_id == current_user.id
    ).group_by(
        extract("month", Expense.expense_date)
    ).all()
    for month, amount in expense_result:
        expense_data[int(month)-1] = float(amount)
    savings_data = []
    for i in range(12):
        savings_data.append(
            income_data[i] - expense_data[i]
        )
    expense_ratio = round(
        (total_expense / total_income) * 100,
        2
    ) if total_income > 0 else 0
    investment_growth = roi
    debt_ratio = 18      
    score = 100
    if savings_ratio < 30:
        score -= 20
    if expense_ratio > 70:
        score -= 20
    if debt_ratio > 40:
        score -= 20
    if roi < 5:
        score -= 10
    score = max(score, 0)
    if score >= 90:
        health_status = "Excellent"
        health_color = "success"
    elif score >= 75:
        health_status = "Good"
        health_color = "primary"
    elif score >= 60:
        health_status = "Fair"
        health_color = "warning"
    else:
        health_status = "Poor"
        health_color = "danger"
    recommendations = []
    if savings_ratio < 30:
        recommendations.append(
            "Increase monthly savings."
        )
    if expense_ratio > 70:
        recommendations.append(
            "Reduce unnecessary expenses."
        )
    if roi < 5:
        recommendations.append(
            "Increase SIP investment."
        )
    if debt_ratio > 40:
        recommendations.append(
            "Reduce outstanding loans."
        )
    if not recommendations:
        recommendations.append(
            "Excellent! Keep maintaining your financial habits."
        )
    notifications = Notification.query.filter_by(
        user_id=current_user.id
    ).order_by(
        Notification.created_at.desc()
    ).limit(5).all()
    budgets = Budget.query.filter_by(
        user_id=current_user.id
    ).all()
    for budget in budgets:
        total_spent = db.session.query(
            func.sum(Expense.amount)
        ).filter(
            Expense.user_id == current_user.id,
            Expense.category == budget.category
        ).scalar() or 0
        if total_spent > budget.budget_amount:
            exists = Notification.query.filter_by(
                user_id=current_user.id,
                title="⚠ Budget Alert",
                message=f"{budget.category} exceeded budget by ₹{total_spent-budget.budget_amount:.2f}"
            ).first()
            if not exists:
                db.session.add(
                    Notification(
                        user_id=current_user.id,
                        title="⚠ Budget Alert",
                        message=f"{budget.category} exceeded budget by ₹{total_spent-budget.budget_amount:.2f}",
                        priority="High",
                        status="Active"
                    )
                )
    if total_savings < 5000:
        exists = Notification.query.filter_by(
            user_id=current_user.id,
            title="🎯 Savings Goal"
        ).first()
        if not exists:
            db.session.add(
                Notification(
                    user_id=current_user.id,
                    title="🎯 Savings Goal",
                    message="Save ₹5,000 more this month to reach your goal.",
                    priority="Medium",
                    status="Active"
                )
            )
    if total_savings < 10000:
        exists = Notification.query.filter_by(
            user_id=current_user.id,
            title="💳 Low Balance"
        ).first()
        if not exists:
            db.session.add(
                Notification(
                    user_id=current_user.id,
                    title="💳 Low Balance",
                    message="Savings account balance below ₹10,000.",
                    priority="High",
                    status="Active"
                )
            )
    if roi > 4:
        exists = Notification.query.filter_by(
            user_id=current_user.id,
            title="📈 Investment Alert"
        ).first()
        if not exists:
            db.session.add(
                Notification(
                    user_id=current_user.id,
                    title="📈 Investment Alert",
                    message=f"Investment portfolio increased by {roi:.2f}%",
                    priority="Low",
                    status="Completed"
                )
            )
    if score < 60:
        exists = Notification.query.filter_by(
            user_id=current_user.id,
            title="📉 Financial Health"
        ).first()
        if not exists:
            db.session.add(
                Notification(
                    user_id=current_user.id,
                    title="📉 Financial Health",
                    message="Your financial health score is below average.",
                    priority="High",
                    status="Active"
                )
            )
    today = date.today()
    if today.day == 25:
        exists = Notification.query.filter_by(
            user_id=current_user.id,
            title="📅 Bill Reminder"
        ).first()
        if not exists:
            db.session.add(
                Notification(
                    user_id=current_user.id,
                    title="📅 Bill Reminder",
                    message="Electricity bill due today.",
                    priority="Medium",
                    status="Pending"
                )
            )
    db.session.commit()
    notifications = Notification.query.filter_by(
        user_id=current_user.id
    ).order_by(
        Notification.created_at.desc()
    ).limit(5).all()
    unread_count = Notification.query.filter_by(
        user_id=current_user.id,
        is_read=False
    ).count()
    spending_analysis = db.session.query(
        Expense.category,
        func.sum(Expense.amount)
    ).filter(
        Expense.user_id == current_user.id
    ).group_by(
        Expense.category
    ).all()
    budget_recommendations = []
    if total_income > 0:
        if savings_ratio < 30:
            budget_recommendations.append(
                "Increase monthly savings by ₹5,000."
            )
        if expense_ratio > 70:
            budget_recommendations.append(
                "Reduce unnecessary expenses."
            )
    budgets = Budget.query.filter_by(
        user_id=current_user.id
    ).all()
    for budget in budgets:
        spent = db.session.query(
            func.sum(Expense.amount)
        ).filter(
            Expense.user_id == current_user.id,
            Expense.category == budget.category
        ).scalar() or 0
        if spent > budget.budget_amount:
            excess = spent - budget.budget_amount
            budget_recommendations.append(
                f"Reduce {budget.category} expenses by ₹{excess:.2f}"
            )
        elif spent >= budget.budget_amount * 0.90:
            budget_recommendations.append(
                f"{budget.category} budget is almost exhausted."
            )
    if len(budget_recommendations) == 0:
        budget_recommendations.append(
            "Excellent! You are managing your budget effectively."
        )
    ai_insights = []
    if savings_ratio >= 30:
        ai_insights.append(
            "✅ Your savings rate is healthy."
        )
    else:
        ai_insights.append(
            "⚠ Increase your monthly savings."
        )
    if expense_ratio > 70:
        ai_insights.append(
            "⚠ Your expenses are consuming most of your income."
        )
    elif expense_ratio < 50:
        ai_insights.append(
            "✅ Excellent expense management."
        )
    if roi >= 10:
        ai_insights.append(
            f"📈 Investment portfolio gained {roi:.2f}%."
        )
    elif roi > 0:
        ai_insights.append(
            f"📊 Portfolio growing steadily ({roi:.2f}%)."
        )
    else:
        ai_insights.append(
            "📉 Investment returns are currently low."
        )
    if total_budget > 0:
        usage = (total_expense / total_budget) * 100
        if usage > 100:
            ai_insights.append(
                "🚨 Budget exceeded."
            )
        elif usage > 90:
            ai_insights.append(
                "⚠ Budget almost exhausted."
            )
        else:
            ai_insights.append(
                "✅ Budget utilization is under control."
            )
    goal_progress = []
    goals = Goal.query.filter_by(
        user_id=current_user.id
    ).all()
    for goal in goals:
        progress = 0
        if goal.target_amount > 0:
            progress = round(
                (goal.saved_amount / goal.target_amount) * 100,
                2
            )
        if progress > 100:
            progress = 100
        goal_progress.append({
            "goal": goal.goal_name,
            "target": goal.target_amount,
            "current": goal.saved_amount,
            "progress": progress
        })
    return render_template(
    "dashboard.html",
    total_income=total_income,
    total_expense=total_expense,
    total_budget=total_budget,
    total_savings=total_savings,
    total_invested=total_invested,
    total_investment=total_investment,
    income_growth=income_growth,
    expense_growth=expense_growth,
    savings_ratio=savings_ratio,
    expense_ratio=expense_ratio,
    investment_growth=investment_growth,
    debt_ratio=debt_ratio,
    roi=roi,
    financial_score=score,
    health_status=health_status,
    health_color=health_color,
    recommendations=recommendations,
    labels=labels,
    values=values,
    months=months,
    income_data=income_data,
    expense_data=expense_data,
    savings_data=savings_data,
    recent_income=recent_income,
    recent_expense=recent_expense,
    recent_budget=recent_budget,
    notifications=notifications,
    unread_count=unread_count,
    spending_analysis=spending_analysis,
    budget_recommendations=budget_recommendations,
    ai_insights=ai_insights,
    goal_progress=goal_progress,
)
    
@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("Logged Out Successfully", "info")
    return redirect(url_for("login"))

@app.route("/expense", methods=["GET","POST"])
@login_required
def expense():
    form = ExpenseForm()
    if form.validate_on_submit():
        expense = Expense(
            user_id=current_user.id,
            category=form.category.data,
            amount=form.amount.data,
            payment_method=form.payment_method.data,
            expense_date=form.expense_date.data,
            description=form.description.data
        )
        db.session.add(expense)
        db.session.commit()
        flash("Expense Added Successfully")
        return redirect(url_for("expense"))
    expenses = Expense.query.filter_by(
        user_id=current_user.id
    ).all()
    return render_template(
        "expense.html",
        form=form,
        expenses=expenses
    )

@app.route("/income", methods=["GET", "POST"])
@login_required
def income():
    form = IncomeForm()
    if form.validate_on_submit():
        data = Income(
            user_id=current_user.id,
            source=form.source.data,
            amount=form.amount.data,
            income_date=form.income_date.data,
            description=form.description.data
        )
        db.session.add(data)
        db.session.commit()
        flash("Income Added Successfully", "success")
        return redirect(url_for("income"))
    incomes = Income.query.filter_by(user_id=current_user.id).all()
    return render_template(
        "income.html",
        form=form,
        incomes=incomes
    )

@app.route("/budget", methods=["GET", "POST"])
@login_required
def budget():
    form = BudgetForm()
    if form.validate_on_submit():
        budget = Budget(
            user_id=current_user.id,
            category=form.category.data,
            budget_amount=form.budget_amount.data,
            month=form.month.data,
            year=form.year.data
        )
        db.session.add(budget)
        db.session.commit()
        flash("Budget Saved Successfully!", "success")
        return redirect(url_for("budget"))
    budgets = Budget.query.filter_by(
        user_id=current_user.id
    ).all()
    return render_template(
        "budget.html",
        form=form,
        budgets=budgets
    )

@app.route("/reports")
@login_required
def reports():
    total_income = db.session.query(
        func.sum(Income.amount)
    ).filter(
        Income.user_id == current_user.id
    ).scalar() or 0
    total_expense = db.session.query(
        func.sum(Expense.amount)
    ).filter(
        Expense.user_id == current_user.id
    ).scalar() or 0
    total_budget = db.session.query(
        func.sum(Budget.budget_amount)
    ).filter(
        Budget.user_id == current_user.id
    ).scalar() or 0
    savings = total_income - total_expense
    income_list = Income.query.filter_by(
        user_id=current_user.id
    ).all()
    expense_list = Expense.query.filter_by(
        user_id=current_user.id
    ).all()
    budget_list = Budget.query.filter_by(
        user_id=current_user.id
    ).all()
    monthly_expenses = db.session.query(
            extract("month", Expense.expense_date),
            func.sum(Expense.amount)
        ).filter(
            Expense.user_id == current_user.id
        ).group_by(
            extract("month", Expense.expense_date)
        ).all()
    monthly_expenses = db.session.query(
        extract("month", Expense.expense_date),
        func.sum(Expense.amount)
    ).filter(
        Expense.user_id == current_user.id
    ).group_by(
        extract("month", Expense.expense_date)
    ).all()
    months = [
        "Jan","Feb","Mar","Apr","May","Jun",
        "Jul","Aug","Sep","Oct","Nov","Dec"
    ]
    expense_data = [0] * 12
    for month, amount in monthly_expenses:
        expense_data[int(month)-1] = float(amount)
    budgets = Budget.query.filter_by(
        user_id=current_user.id
    ).all()
    budget_report = []
    for budget in budgets:
        spent = db.session.query(
            func.sum(Expense.amount)
        ).filter_by(
            user_id=current_user.id,
            category=budget.category
        ).scalar() or 0
        utilization = min(
        (spent / budget.budget_amount) * 100,
        100
        ) if budget.budget_amount else 0
        budget_report.append({
            "category": budget.category,
            "budget": budget.budget_amount,
            "spent": spent,
            "utilization": round(utilization,2)
        })
    investments = Investment.query.filter_by(
        user_id=current_user.id
    ).all()
    investment_report = []
    for inv in investments:
        profit = inv.current_value - inv.amount
        roi = (
            profit / inv.amount
        ) * 100 if inv.amount else 0
        investment_report.append({
            "investment": inv.investment_name,
            "invested": inv.amount,
            "current": inv.current_value,
            "profit": profit,
            "roi": round(roi,2)
        })
    goals = Goal.query.filter_by(
        user_id=current_user.id
    ).all()
    goal_report = []
    for goal in goals:
        progress = (
            goal.saved_amount /
            goal.target_amount
        ) * 100 if goal.target_amount else 0
        goal_report.append({
            "goal": goal.goal_name,
            "target": goal.target_amount,
            "current": goal.saved_amount,
            "progress": round(progress,2)
        })
    return render_template(
        "reports.html",
        total_income=total_income,
        total_expense=total_expense,
        total_budget=total_budget,
        savings=savings,
        income_list=income_list,
        expense_list=expense_list,
        budget_list=budget_list,
        months=months,
        expense_data=expense_data,
        budget_report=budget_report,
        investment_report=investment_report,
        goal_report=goal_report        
    )

@app.route("/investment", methods=["GET", "POST"])
@login_required
def investment():
    form = InvestmentForm()
    if form.validate_on_submit():
        investment = Investment(
            user_id=current_user.id,
            investment_type=form.investment_type.data,
            investment_name=form.investment_name.data,
            amount=form.amount.data,
            current_value=form.current_value.data,
            investment_date=form.investment_date.data
        )
        db.session.add(investment)
        db.session.commit()
        flash("Investment Added Successfully!", "success")
        return redirect(url_for("investment"))
    investments = Investment.query.filter_by(
        user_id=current_user.id
    ).all()
    total_investment = sum(float(i.amount or 0) for i in investments)
    total_current = sum(float(i.current_value or 0) for i in investments)
    total_profit = total_current - total_investment
    roi = round(
        (total_profit / total_investment) * 100, 2
    ) if total_investment > 0 else 0
    chart_data = db.session.query(
        Investment.investment_type,
        func.sum(Investment.current_value)
    ).filter(
        Investment.user_id == current_user.id
    ).group_by(
        Investment.investment_type
    ).all()
    labels = []
    values = []
    percentages = []
    for asset_type, amount in chart_data:
        labels.append(asset_type)
        amount = float(amount)
        values.append(amount)
        percentages.append(
            round((amount / total_current) * 100, 2)
            if total_current > 0 else 0
        )
    investment_names = [
        inv.investment_name for inv in investments
    ]
    invested_amounts = [
        float(inv.amount) for inv in investments
    ]
    current_values = [
        float(inv.current_value) for inv in investments
    ]
    profits = [
        float(inv.current_value - inv.amount)
        for inv in investments
    ]
    top_asset = None
    if investments:
        top_asset = max(
            investments,
            key=lambda x: x.current_value - x.amount
        )
    diversification = []
    for inv in investments:
        profit = float(inv.current_value - inv.amount)
        investment_roi = (
            round((profit / inv.amount) * 100, 2)
            if inv.amount > 0 else 0
        )
        allocation = (
            round((inv.current_value / total_current) * 100, 2)
            if total_current > 0 else 0
        )
        diversification.append({
            "name": inv.investment_name,
            "type": inv.investment_type,
            "invested": float(inv.amount),
            "current": float(inv.current_value),
            "profit": profit,
            "roi": investment_roi,
            "allocation": allocation
        })
    return render_template(
        "investment.html",
        form=form,
        investments=investments,
        total_investment=total_investment,
        total_current=total_current,
        total_profit=total_profit,
        roi=roi,
        labels=labels,
        values=values,
        percentages=percentages,
        investment_names=investment_names,
        invested_amounts=invested_amounts,
        current_values=current_values,
        profits=profits,
        diversification=diversification,
        top_asset=top_asset
    )

@app.route("/goal", methods=["GET", "POST"])
@login_required
def goal():
    form = GoalForm()
    if form.validate_on_submit():
        new_goal = Goal(
            user_id=current_user.id,
            goal_name=form.goal_name.data,
            target_amount=form.target_amount.data,
            saved_amount=form.saved_amount.data,
            target_date=form.target_date.data
        )
        db.session.add(new_goal)
        db.session.commit()
        flash("Financial Goal Added Successfully!", "success")
        return redirect(url_for("goal"))
    goals = Goal.query.filter_by(
        user_id=current_user.id
    ).order_by(
        Goal.target_date.asc()
    ).all()
    total_target = sum(float(g.target_amount or 0) for g in goals)
    total_saved = sum(float(g.saved_amount or 0) for g in goals)
    remaining_amount = total_target - total_saved
    if total_target > 0:
        overall_progress = round(
            (total_saved / total_target) * 100,
            2
        )
    else:
        overall_progress = 0
    goal_labels = []
    goal_progress = []
    for g in goals:
        goal_labels.append(g.goal_name)
        if g.target_amount > 0:
            progress = round(
                (g.saved_amount / g.target_amount) * 100,
                2
            )
        else:
            progress = 0
        goal_progress.append(progress)
    top_goal = None
    if goals:
        top_goal = min(
            goals,
            key=lambda x: x.target_date
        )
    analytics = []
    for g in goals:
        saved = float(g.saved_amount or 0)
        target = float(g.target_amount or 0)
        remaining = target - saved
        if target > 0:
            progress = round(
                (saved / target) * 100,
                2
            )
        else:
            progress = 0
        if progress >= 100:
            status = "Completed"
        elif progress >= 75:
            status = "Almost There"
        elif progress >= 40:
            status = "In Progress"
        else:
            status = "Started"
        analytics.append({
            "goal_name": g.goal_name,
            "target": target,
            "saved": saved,
            "remaining": remaining,
            "progress": progress,
            "status": status,
            "target_date": g.target_date
        })
    return render_template(
        "goal.html",
        form=form,
        goals=goals,
        goal_labels=goal_labels,
        goal_progress=goal_progress,
        total_target=total_target,
        total_saved=total_saved,
        remaining_amount=remaining_amount,
        overall_progress=overall_progress,
        top_goal=top_goal,
        analytics=analytics
    )

@app.route("/portfolio")
@login_required
def portfolio():
    investments = Investment.query.filter_by(
        user_id=current_user.id
    ).all()
    goals = Goal.query.filter_by(
        user_id=current_user.id
    ).all()
    total_goals = len(goals)
    completed = len([
        g for g in goals
        if g.saved_amount >= g.target_amount
    ])
    goal_percentage = round(
        (completed / total_goals) * 100,
        2
    ) if total_goals else 0
    total_investment = sum(float(i.amount) for i in investments)
    total_current = sum(float(i.current_value) for i in investments)
    total_profit = total_current - total_investment
    roi = round(
        (total_profit / total_investment) * 100,
        2
    ) if total_investment else 0
    asset_data = db.session.query(
        Investment.investment_type,
        func.sum(Investment.current_value)
    ).filter(
        Investment.user_id == current_user.id
    ).group_by(
        Investment.investment_type
    ).all()
    asset_labels = []
    asset_values = []
    for name, value in asset_data:
        asset_labels.append(name)
        asset_values.append(float(value))
        monthly = db.session.query(
        func.date_format(
            Investment.investment_date,
            "%Y-%m"
            ),
        func.sum(Investment.current_value)
        ).filter(
        Investment.user_id == current_user.id
        ).group_by(
        func.date_format(
            Investment.investment_date,
            "%Y-%m"
        )
        ).all()
    months = []
    growth = []
    for month, value in monthly:
        months.append(month)
        growth.append(float(value))
        top_asset = None
        worst_asset = None
    if investments:
        top_asset = max(
            investments,
            key=lambda x: x.current_value - x.amount
        )
        worst_asset = min(
            investments,
            key=lambda x: x.current_value - x.amount
        )
        total_goals = len(goals)
    completed = len([
        g for g in goals
        if g.saved_amount >= g.target_amount
    ])
    goal_percentage = round(
        (completed / total_goals) * 100,
        2
    ) if total_goals else 0
    risk = "Low"
    crypto = 0
    stocks = 0
    for i in investments:
        if i.investment_type == "Cryptocurrency":
            crypto += 1
        if i.investment_type == "Stocks":
            stocks += 1
    if crypto >= 2:
        risk = "High"
    elif stocks >= 2:
        risk = "Medium"
    diversification = []
    for inv in investments:
        allocation = round(
            (float(inv.current_value) / total_current) * 100,
            2
        ) if total_current else 0
        diversification.append({
            "name": inv.investment_name,
            "type": inv.investment_type,
            "allocation": allocation
        })
        return render_template(
        "portfolio.html",
        total_investment=total_investment,
        total_current=total_current,
        total_profit=total_profit,
        roi=roi,
        asset_labels=asset_labels,
        asset_values=asset_values,
        months=months,
        growth=growth,
        diversification=diversification,
        top_asset=top_asset,
        worst_asset=worst_asset,
        total_goals=total_goals,
        completed=completed,
        goal_percentage=goal_percentage,
        risk=risk
    )

@app.route("/download_pdf")
@login_required
def download_pdf():
    investments = Investment.query.filter_by(
        user_id=current_user.id
    ).all()
    total_investment = sum(float(i.amount) for i in investments)
    total_current = sum(float(i.current_value) for i in investments)
    total_profit = total_current - total_investment
    roi = round(
        (total_profit / total_investment) * 100,
        2
    ) if total_investment else 0
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer)
    styles = getSampleStyleSheet()
    elements = []
    elements.append(
        Paragraph(
            "<b>Smart Finance Insights</b>",
            styles["Title"]
        )
    )
    elements.append(
        Paragraph(
            "Portfolio Analytics Report",
            styles["Heading2"]
        )
    )
    elements.append(
        Paragraph("<br/>", styles["BodyText"])
    )
    summary = [
        ["Total Investment", f"₹ {total_investment:,.2f}"],
        ["Current Value", f"₹ {total_current:,.2f}"],
        ["Profit / Loss", f"₹ {total_profit:,.2f}"],
        ["Overall ROI", f"{roi}%"]
    ]
    table = Table(summary, colWidths=[3*inch,2.5*inch])
    table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.grey),
        ("GRID",(0,0),(-1,-1),1,colors.black),
        ("BACKGROUND",(0,0),(0,-1),colors.lightgrey),
        ("TEXTCOLOR",(0,0),(-1,-1),colors.black),
        ("BOTTOMPADDING",(0,0),(-1,-1),8)
    ]))
    elements.append(table)
    elements.append(
        Paragraph("<br/><b>Investment Details</b>", styles["Heading2"])
    )
    data = [[
        "Type",
        "Name",
        "Investment",
        "Current",
        "Profit"
    ]]
    for inv in investments:
        data.append([
            inv.investment_type,
            inv.investment_name,
            f"₹ {inv.amount}",
            f"₹ {inv.current_value}",
            f"₹ {inv.current_value-inv.amount}"
        ])
    investment_table = Table(data)
    investment_table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.darkblue),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("GRID",(0,0),(-1,-1),1,colors.black),
        ("BACKGROUND",(0,1),(-1,-1),colors.beige)
    ]))
    elements.append(investment_table)
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response = make_response(pdf)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = "attachment; filename=Portfolio_Report.pdf"
    return response

@app.route("/download_excel")
@login_required
def download_excel():
    investments = Investment.query.filter_by(
        user_id=current_user.id
    ).all()
    total_investment = sum(float(i.amount) for i in investments)
    total_current = sum(float(i.current_value) for i in investments)
    total_profit = total_current - total_investment
    roi = round(
        (total_profit / total_investment) * 100,
        2
    ) if total_investment > 0 else 0
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Report"
    ws["A1"] = "Smart Finance Insights"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A2"] = "Portfolio Analytics Report"
    ws["A2"].font = Font(size=14, bold=True)
    ws["A4"] = "Total Investment"
    ws["B4"] = total_investment
    ws["A5"] = "Current Value"
    ws["B5"] = total_current
    ws["A6"] = "Profit / Loss"
    ws["B6"] = total_profit
    ws["A7"] = "Overall ROI (%)"
    ws["B7"] = roi
    row = 10
    headings = [
        "Investment Type",
        "Investment Name",
        "Investment Amount",
        "Current Value",
        "Profit / Loss",
        "ROI (%)"
    ]
    fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )
    for col, heading in enumerate(headings, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = heading
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    row += 1
    for inv in investments:
        profit = float(inv.current_value - inv.amount)
        investment_roi = round(
            (profit / inv.amount) * 100,
            2
        ) if inv.amount > 0 else 0
        ws.cell(row=row, column=1).value = inv.investment_type
        ws.cell(row=row, column=2).value = inv.investment_name
        ws.cell(row=row, column=3).value = float(inv.amount)
        ws.cell(row=row, column=4).value = float(inv.current_value)
        ws.cell(row=row, column=5).value = profit
        ws.cell(row=row, column=6).value = investment_roi
        row += 1
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 5
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers[
        "Content-Disposition"
    ] = "attachment; filename=Portfolio_Report.xlsx"
    response.headers[
        "Content-Type"
    ] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response

@app.route("/spending_analysis")
@login_required
def spending_analysis():
    total_expense = db.session.query(
        func.sum(Expense.amount)
    ).filter(
        Expense.user_id == current_user.id
    ).scalar() or 0
    total_budget = db.session.query(
        func.sum(Budget.budget_amount)
    ).filter(
        Budget.user_id == current_user.id
    ).scalar() or 0
    remaining_budget = total_budget - total_expense
    category_data = db.session.query(
        Expense.category,
        func.sum(Expense.amount)
    ).filter(
        Expense.user_id == current_user.id
    ).group_by(
        Expense.category
    ).all()
    category_summary = []
    category_labels = []
    category_values = []
    highest_category = "N/A"
    highest_amount = 0
    highest_percentage = 0
    for category, amount in category_data:
        amount = float(amount)
        percentage = 0
        if total_expense > 0:
            percentage = round(
                amount / total_expense * 100,
                2
            )
        category_summary.append({
            "category": category,
            "amount": amount,
            "percentage": percentage
        })
        category_labels.append(category)
        category_values.append(amount)
        if amount > highest_amount:
            highest_amount = amount
            highest_category = category
            highest_percentage = percentage
    months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    monthly_expenses = [0] * 12
    monthly_result = db.session.query(
        extract("month", Expense.expense_date),
        func.sum(Expense.amount)
    ).filter(
        Expense.user_id == current_user.id
    ).group_by(
        extract("month", Expense.expense_date)
    ).all()
    for month, amount in monthly_result:
        monthly_expenses[int(month) - 1] = float(amount)
    budget_utilization = 0
    if total_budget > 0:
        budget_utilization = round(
            total_expense /
            total_budget * 100,
            2
        )
    if budget_utilization >= 100:
        recommendation = (
            "Your expenses exceeded your monthly budget. "
            "Reduce unnecessary spending."
        )
    elif budget_utilization >= 80:
        recommendation = (
            "You have used most of your budget. "
            "Spend carefully for the rest of the month."
        )
    elif budget_utilization >= 50:
        recommendation = (
            "Your spending is under control. "
            "Continue maintaining your budget."
        )
    else:
        recommendation = (
            "Excellent! Your spending is well below your budget."
        )
    recent_expenses = Expense.query.filter_by(
        user_id=current_user.id
    ).order_by(
        Expense.expense_date.desc()
    ).limit(10).all()
    return render_template(
        "spending_analysis.html",
        total_expense=total_expense,
        total_budget=total_budget,
        remaining_budget=remaining_budget,
        highest_category=highest_category,
        highest_percentage=highest_percentage,
        category_summary=category_summary,
        category_labels=category_labels,
        category_values=category_values,
        months=months,
        monthly_expenses=monthly_expenses,
        budget_utilization=budget_utilization,
        recommendation=recommendation,
        recent_expenses=recent_expenses
    )

@app.route("/budget_recommendation")
@login_required
def budget_recommendation():
    total_expense = db.session.query(
        func.sum(Expense.amount)
    ).filter(
        Expense.user_id == current_user.id
    ).scalar() or 0
    total_budget = db.session.query(
        func.sum(Budget.budget_amount)
    ).filter(
        Budget.user_id == current_user.id
    ).scalar() or 0
    savings = total_budget - total_expense
    expense_result = db.session.query(
        Expense.category,
        func.sum(Expense.amount)
    ).filter(
        Expense.user_id == current_user.id
    ).group_by(
        Expense.category
    ).all()
    expense_summary = {}
    for category, amount in expense_result:
        expense_summary[category] = float(amount)
    budget_result = db.session.query(
        Budget.category,
        func.sum(Budget.budget_amount)
    ).filter(
        Budget.user_id == current_user.id
    ).group_by(
        Budget.category
    ).all()
    budget_summary = {}
    for category, amount in budget_result:
        budget_summary[category] = float(amount)
    recommendations = []
    overspending_count = 0
    for category, expense in expense_summary.items():
        budget = budget_summary.get(category, 0)
        difference = expense - budget
        if budget == 0:
            recommended_budget = expense * 1.10
            status = "No Budget"
            recommendation = "Create a monthly budget."
            alert = "Warning"
        elif expense > budget:
            overspending_count += 1
            recommended_budget = expense * 0.90
            status = "Overspending"
            recommendation = (
                f"Reduce spending by ₹{difference:,.2f}"
            )
            alert = "High"
        else:
            recommended_budget = budget
            status = "Good"
            recommendation = "Maintain current spending."
            alert = "Normal"
        recommendations.append({
            "category": category,
            "expense": expense,
            "budget": budget,
            "recommended_budget": recommended_budget,
            "difference": difference,
            "status": status,
            "alert": alert,
            "recommendation": recommendation
        })
    suggestions = []
    for rec in recommendations:
        if rec["status"] == "Overspending":
            suggestions.append(
                f"Reduce {rec['category']} spending by "
                f"₹{rec['difference']:.2f}"
            )
    if savings < 5000:
        suggestions.append(
            "Increase monthly savings by ₹5,000."
        )
    if len(suggestions) == 0:
        suggestions.append(
            "Excellent! Your spending is under control."
        )
    chart_labels = []
    expense_values = []
    budget_values = []
    for rec in recommendations:
        chart_labels.append(rec["category"])
        expense_values.append(rec["expense"])
        budget_values.append(rec["budget"])
    return render_template(
        "budget_recommendation.html",
        total_expense=total_expense,
        total_budget=total_budget,
        savings=savings,
        overspending_count=overspending_count,
        recommendations=recommendations,
        suggestions=suggestions,
        chart_labels=chart_labels,
        expense_values=expense_values,
        budget_values=budget_values
    )



def create_notification(user_id,title,message,priority="Medium",status="Active"):
    notification = Notification(
        user_id=user_id,
        title=title,
        message=message,
        priority=priority,
        status=status,
        is_read=False,
        created_at=datetime.utcnow()
    )
    db.session.add(notification)
    db.session.commit()
    return notification

@app.route("/notification/read/<int:id>", methods=["POST"])
@login_required
def mark_notification_read(id):
    notification = Notification.query.filter_by(
        id=id,
        user_id=current_user.id
    ).first()
    if notification is None:
        return jsonify({
            "success": False,
            "message": "Notification not found."
        }), 404
    notification.is_read = True
    db.session.commit()
    unread_count = Notification.query.filter_by(
        user_id=current_user.id,
        is_read=False
    ).count()
    return jsonify({
        "success": True,
        "message": "Notification marked as read.",
        "unread_count": unread_count,
        "notification_id": notification.id
    })

@app.route("/delete_notification/<int:id>", methods=["POST"])
@login_required
def delete_notification(id):
    notification = Notification.query.filter_by(
        id=id,
        user_id=current_user.id
    ).first()
    if notification is None:
        return jsonify({
            "success": False,
            "message": "Notification not found."
        }), 404
    db.session.delete(notification)
    db.session.commit()
    unread_count = Notification.query.filter_by(
        user_id=current_user.id,
        is_read=False
    ).count()
    return jsonify({
        "success": True,
        "message": "Notification deleted successfully.",
        "unread_count": unread_count,
        "notification_id": id
    })
    
@app.route("/clear_notifications", methods=["POST"])
@login_required
def clear_notifications():
    deleted_count = Notification.query.filter_by(
        user_id=current_user.id
    ).delete(synchronize_session=False)
    db.session.commit()
    return jsonify({
        "success": True,
        "message": f"{deleted_count} notifications cleared.",
        "unread_count": 0
    })


@app.context_processor
def inject_notifications():
    if current_user.is_authenticated:
        notifications = Notification.query.filter_by(
            user_id=current_user.id
        ).order_by(
            Notification.created_at.desc()
        ).all()
        unread_count = Notification.query.filter_by(
            user_id=current_user.id,
            is_read=False
        ).count()
    else:
        notifications = []
        unread_count = 0
    return {
        "notifications": notifications,
        "unread_count": unread_count
    }

@app.route("/ai_insights")
@login_required
def ai_insights():
    total_income = db.session.query(
        func.sum(Income.amount)
    ).filter(
        Income.user_id == current_user.id
    ).scalar() or 0
    total_expense = db.session.query(
        func.sum(Expense.amount)
    ).filter(
        Expense.user_id == current_user.id
    ).scalar() or 0
    total_budget = db.session.query(
        func.sum(Budget.budget_amount)
    ).filter(
        Budget.user_id == current_user.id
    ).scalar() or 0
    total_investment = db.session.query(
        func.sum(Investment.current_value)
    ).filter(
        Investment.user_id == current_user.id
    ).scalar() or 0
    total_invested = db.session.query(
        func.sum(Investment.amount)
    ).filter(
        Investment.user_id == current_user.id
    ).scalar() or 0
    savings = total_income - total_expense
    savings_ratio = round(
        (savings / total_income) * 100,
        2
    ) if total_income > 0 else 0
    expense_ratio = round(
        (total_expense / total_income) * 100,
        2
    ) if total_income > 0 else 0
    budget_utilization = round(
        (total_expense / total_budget) * 100,
        2
    ) if total_budget > 0 else 0
    roi = round(
        ((total_investment - total_invested) / total_invested) * 100,
        2
    ) if total_invested > 0 else 0
    recommendations = []
    if savings_ratio >= 30:
        recommendations.append({
            "type": "success",
            "title": "Excellent Savings",
            "message": "Your savings ratio is excellent. Keep maintaining this habit."
        })
    elif savings_ratio >= 20:
        recommendations.append({
            "type": "info",
            "title": "Good Savings",
            "message": "Try increasing your monthly savings by ₹2,000."
        })
    else:
        recommendations.append({
            "type": "danger",
            "title": "Low Savings",
            "message": "Increase your monthly savings by at least ₹5,000."
        })
    if expense_ratio > 80:
        recommendations.append({
            "type": "danger",
            "title": "High Expenses",
            "message": "Your expenses are too high. Reduce unnecessary spending."
        })
    elif expense_ratio > 60:
        recommendations.append({
            "type": "warning",
            "title": "Moderate Expenses",
            "message": "Monitor discretionary spending."
        })
    else:
        recommendations.append({
            "type": "success",
            "title": "Healthy Expenses",
            "message": "Excellent expense management."
        })
    if budget_utilization > 100:
        recommendations.append({
            "type": "danger",
            "title": "Budget Exceeded",
            "message": "You have exceeded your monthly budget."
        })
    elif budget_utilization > 90:
        recommendations.append({
            "type": "warning",
            "title": "Budget Alert",
            "message": "You are close to reaching your budget limit."
        })
    else:
        recommendations.append({
            "type": "success",
            "title": "Budget Under Control",
            "message": "Great job managing your budget."
        })
    if roi > 15:
        recommendations.append({
            "type": "success",
            "title": "Investment Growth",
            "message": "Your investments are performing very well."
        })
    elif roi >= 8:
        recommendations.append({
            "type": "info",
            "title": "Investment Suggestion",
            "message": "Increase SIP investment by ₹2,000 every month."
        })
    else:
        recommendations.append({
            "type": "warning",
            "title": "Improve Investments",
            "message": "Diversify your investments into Mutual Funds or ETFs."
        })
    if savings < total_expense * 6:
        recommendations.append({
            "type": "warning",
            "title": "Emergency Fund",
            "message": "Build an emergency fund covering at least 6 months of expenses."
        })
    else:
        recommendations.append({
            "type": "success",
            "title": "Emergency Fund",
            "message": "Your emergency fund looks healthy."
        })
    recommendations.append({
        "type": "primary",
        "title": "Smart Tip",
        "message": "Review subscriptions regularly to eliminate unnecessary recurring expenses."
    })
    recommendations.append({
        "type": "primary",
        "title": "Investment Tip",
        "message": "Invest consistently every month instead of timing the market."
    })
    recommendations.append({
        "type": "primary",
        "title": "Savings Tip",
        "message": "Maintain savings above 30% of your monthly income."
    })
    months = [
        "Jan","Feb","Mar","Apr","May","Jun",
        "Jul","Aug","Sep","Oct","Nov","Dec"
    ]
    income_data = [0] * 12
    expense_data = [0] * 12
    income_result = db.session.query(
        extract("month", Income.income_date),
        func.sum(Income.amount)
        ).filter(
        Income.user_id == current_user.id
        ).group_by(
        extract("month", Income.income_date)
        ).all()
    for month, amount in income_result:
        income_data[int(month)-1] = float(amount)
    expense_result = db.session.query(
        extract("month", Expense.expense_date),
        func.sum(Expense.amount)
        ).filter(
        Expense.user_id == current_user.id
        ).group_by(
        extract("month", Expense.expense_date)
        ).all()
    for month, amount in expense_result:
        expense_data[int(month)-1] = float(amount)
    expense_chart = db.session.query(
        Expense.category,
        func.sum(Expense.amount)
        ).filter(
        Expense.user_id == current_user.id
        ).group_by(
        Expense.category
        ).all()
    labels = []
    values = []
    for category, amount in expense_chart:
        labels.append(category)
        values.append(float(amount))
    savings_data = []
    for i in range(12):
        savings_data.append(
            income_data[i] - expense_data[i]
        )
    health_score = 0
    health_score += savings_ratio * 0.40
    health_score += roi * 0.20
    health_score += (100 - expense_ratio) * 0.25
    health_score += (100 - budget_utilization) * 0.15
    health_score = round(min(100, health_score))
    if health_score >= 90:
        health_status = "Excellent"
    elif health_score >= 75:
        health_status = "Good"
    elif health_score >= 60:
        health_status = "Fair"
    else:
        health_status = "Poor"
    return render_template(
    "ai_insights.html",
    total_income=total_income,
    total_expense=total_expense,
    total_budget=total_budget,
    total_investment=total_investment,
    total_invested=total_invested,
    savings=savings,
    savings_ratio=savings_ratio,
    expense_ratio=expense_ratio,
    budget_utilization=budget_utilization,
    roi=roi,
    recommendations=recommendations,
    months=months,
    income_data=income_data,
    expense_data=expense_data,
    labels=labels,
    values=values,
    savings_data=savings_data,
    health_score=health_score,
    health_status=health_status
)




@app.route("/export_expense_pdf")
@login_required
def export_expense_pdf():
    expenses = Expense.query.filter_by(
        user_id=current_user.id
    ).all()
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer)
    elements = []
    styles = getSampleStyleSheet()
    elements.append(
        Paragraph("Expense Report", styles["Heading1"])
    )
    data = [
        ["ID","Category","Amount","Date"]
    ]
    for e in expenses:
        data.append([
            e.expense_id,
            e.category,
            f"₹{e.amount}",
            str(e.expense_date)
        ])
    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.grey),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("GRID",(0,0),(-1,-1),1,colors.black),
        ("BACKGROUND",(0,1),(-1,-1),colors.beige),
        ("ALIGN",(0,0),(-1,-1),"CENTER")
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="Expense_Report.pdf",
        mimetype="application/pdf"
    )


@app.route("/export_investment_excel")
@login_required
def export_investment_excel():
    investments = Investment.query.filter_by(
        user_id=current_user.id
    ).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Investment Report"
    ws.append([
        "Investment",
        "Amount Invested",
        "Current Value",
        "Profit",
        "ROI %"
    ])
    for inv in investments:
        profit = inv.current_value - inv.amount
        roi = 0
        if inv.amount > 0:
            roi = round(
                (profit/inv.amount)*100,
                2
            )
        ws.append([
            inv.investment_name,
            inv.amount,
            inv.current_value,
            profit,
            roi
        ])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name="Investment_Report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/jarvis")
@login_required
def jarvis():
    return render_template("jarvis.html")

@app.route("/jarvis/chat", methods=["POST"])
@login_required
def jarvis_chat():
    data = request.get_json()
    message = data.get("message", "").lower().strip()
    total_income = db.session.query(
        func.sum(Income.amount)
    ).filter(
        Income.user_id == current_user.id
    ).scalar() or 0

    monthly_income = db.session.query(
        func.sum(Income.amount)
    ).filter(
        Income.user_id==current_user.id,
        extract("month",Income.income_date)==date.today().month,
        extract("year",Income.income_date)==date.today().year
    ).scalar() or 0

    total_expense = db.session.query(
        func.sum(Expense.amount)
    ).filter(
        Expense.user_id==current_user.id
    ).scalar() or 0

    monthly_expense = db.session.query(
        func.sum(Expense.amount)
    ).filter(
        Expense.user_id==current_user.id,
        extract("month",Expense.expense_date)==date.today().month,
        extract("year",Expense.expense_date)==date.today().year
    ).scalar() or 0
    savings = total_income-total_expense

    total_budget=db.session.query(
    func.sum(Budget.budget_amount)
    ).filter(
        Budget.user_id==current_user.id
    ).scalar() or 0
    budget_remaining=total_budget-total_expense

    budget_utilization=0
    goals = Goal.query.filter_by(
        user_id=current_user.id
    ).all()
    goal_report = []
    total_target = 0
    total_saved = 0
    for goal in goals:
        progress = 0
        if goal.target_amount > 0:
            progress = min(
                round((goal.saved_amount / goal.target_amount) * 100, 2),
                100
            )
        remaining = max(
            goal.target_amount - goal.saved_amount,
            0
        )
        status = "Completed"
        if progress < 100:
            status = "In Progress"
        goal_report.append({
            "goal": goal.goal_name,
            "target": goal.target_amount,
            "saved": goal.saved_amount,
            "remaining": remaining,
            "progress": progress,
            "status": status,
            "date": goal.target_date
        })
        total_target += goal.target_amount
        total_saved += goal.saved_amount
        if total_budget>0:
            budget_utilization=round(
                (total_expense/total_budget)*100,
                2
            )
    overall_progress = 0
    if total_target > 0:
        overall_progress = round(
            (total_saved / total_target) * 100,
            2
        )

    invested = db.session.query(
        func.sum(Investment.amount)
    ).filter(
        Investment.user_id == current_user.id
    ).scalar() or 0

    current_value = db.session.query(
        func.sum(Investment.current_value)
    ).filter(
        Investment.user_id == current_user.id
    ).scalar() or 0

    investment_profit = current_value - invested

    goals = Goal.query.filter_by(
        user_id=current_user.id
    ).all()

    goal_progress = []
    for goal in goals:
        progress = 0
        if goal.target_amount > 0:
            progress = round(
                (goal.saved_amount / goal.target_amount) * 100,
                2
            )

    goal_progress.append({
        "goal": goal.goal_name,
        "saved": goal.saved_amount,
        "target": goal.target_amount,
        "progress": progress
    })
    roi = 0
    if invested > 0:
        roi = round((investment_profit / invested) * 100, 2)
        financial_summary = {
            "income": total_income,
            "expense": total_expense,
            "saving": savings,
            "budget": total_budget,
            "investment": current_value,
            "goal": goal_progress
        }

    investments = Investment.query.filter_by(
        user_id=current_user.id
    ).all()

    total_invested = 0
    total_current = 0

    investment_report = []
    for inv in investments:
        profit = (inv.current_value or 0) - (inv.amount or 0)
        roi = 0
        if inv.amount and inv.amount > 0:
            roi = round((profit / inv.amount) * 100, 2)
        investment_report.append({
            "name": inv.investment_name,
            "type": inv.investment_type,
            "invested": inv.amount,
            "current": inv.current_value,
            "profit": profit,
            "roi": roi
        })
        total_invested += inv.amount or 0
        total_current += inv.current_value or 0
    total_profit = total_current - total_invested
    overall_roi = 0
    if total_invested > 0:
        overall_roi = round((total_profit / total_invested) * 100, 2)
    financial_summary = {
        "income": total_income,
        "expense": total_expense,
        "savings": savings,
        "budget": total_budget,
        "remaining_budget": budget_remaining,
        "budget_utilization": budget_utilization,
        "invested": invested,
        "current_value": current_value,
        "investment_profit": investment_profit,
        "monthly_income": monthly_income,
        "monthly_expense": monthly_expense,
        "goals": goal_progress
    }
    if "monthly income" in message:
        reply = f"📅 Monthly Income : ₹{monthly_income:,.2f}"
    elif "total income" in message:
        reply = f"💰 Total Income : ₹{total_income:,.2f}"
    elif "expense" == message or "total expense" in message:
        reply=f"💸 Total Expense : ₹{total_expense:,.2f}"
    elif "monthly expense" in message:
        reply=f"📅 Monthly Expense : ₹{monthly_expense:,.2f}"
    elif "saving" in message:
        reply=f"💵 Total Savings : ₹{savings:,.2f}"
    elif any(word in message for word in ["hello","hi","hey"]):
            reply = f"""
            👋 Hello {current_user.fullname}!<br/>
            I am JARVIS AI Financial Assistant.<br/>
            How can I help you today?</br/>
            """
    elif "help" in message:
            reply = """
                Available Commands<br/>
                __________________________<br/>
                monthly income<br/>
                total income<br/>
                expense<br/>
                monthly expense<br/>
                total expense<br/>
                savings<br/>
                budget<br/>
                budget utilization<br/>
                investment report<br/>
                goal reports<br/>
                completed goals<br/>
                incompleted goals<br/>
                financial summary<br/>
                income vs expense<br/>
            """
    elif "income vs expense" in message:
        reply=f"""
        Income : ₹{total_income:,.2f}<br/>
        Expense : ₹{total_expense:,.2f}<br/>
        Savings : ₹{savings:,.2f}<br/>
       """
    elif "budget"==message:
        reply=f"""
        Budget : ₹{total_budget:,.2f}<br/>
        Remaining : ₹{budget_remaining:,.2f}<br/>
        """
    elif "budget utilization" in message:
        reply=f"""
        Budget Used : {budget_utilization}%<br/>
        Remaining : ₹{budget_remaining:,.2f}<br/>
        """
    elif "goal report" in message:
        if goal_report:
            reply = (
                "🎯 GOAL REPORT<br/>"
                f"Total Target : ₹{total_target:,.2f}<br/>"
                f"Total Saved : ₹{total_saved:,.2f}<br/>"
                f"Overall Progress : {overall_progress}%<br/>"
            )
            for g in goal_report:
                reply += (
                    f"🎯 {g['goal']}<br>"
                    f"Target : ₹{g['target']:,.2f}<br/>"
                    f"Saved : ₹{g['saved']:,.2f}</br>"
                    f"Remaining : ₹{g['remaining']:,.2f}<br/>"
                    f"Progress : {g['progress']}%<br/>"
                    f"Status : {g['status']}<br/>"
                    f"Target Date : {g['date']}<br/><br/>"
                )
        else:
            reply = "No financial goals found."
    elif "incompleted goals" in message:
                incompleted = [
                    g for g in goal_report
                    if g["progress"] < 100
                ]
                if incompleted:
                    reply = "🏆 Incompleted Goals<br/>"
                    for g in incompleted:
                        reply += f" ❌{g['goal']}<br/>"
                else:
                    reply = "No incompleted goals."
    elif "completed goals" in message:
        completed = [
            g for g in goal_report
            if g["progress"] >= 100
        ]
        if completed:
            reply = "🏆 Completed Goals<br/>"
            for g in completed:
                reply += f"✅ {g['goal']}<br/>"
        else:
            reply = "No completed goals."
    elif "investment report" in message:
        if investment_report:
            reply = (
                "📈 Investment Report<br/>"
                f"💰 Total Invested : ₹{total_invested:,.2f}<br/>"
                f"💵 Current Value : ₹{total_current:,.2f}<br/>"
                f"📈 Total Profit : ₹{total_profit:,.2f}<br/>"
                f"📊 Overall ROI : {overall_roi}%<br/>"
                "-------------------------<br/>"
            )
            for inv in investment_report:
                reply += (
                    f"📌 {inv['name']}<br/>"
                    f"Type : {inv['type']}<br/>"
                    f"Invested : ₹{inv['invested']:,.2f}<br/>"
                    f"Current : ₹{inv['current']:,.2f}<br/>"
                    f"Profit : ₹{inv['profit']:,.2f}<br/>"
                    f"ROI : {inv['roi']}%<br/>"
                )
        else:
            reply = "No investment records found."
    elif (
        "financial summary" in message
    ):
        reply = (
            "📊 FINANCIAL SUMMARY<br/>"
            f"💰 Total Income<br/>"
            f"₹{financial_summary['income']:,.2f}<br/>"
            f"💸 Total Expense<br/>"
            f"₹{financial_summary['expense']:,.2f}<br/>"
            f"💵 Savings<br/>"
            f"₹{financial_summary['savings']:,.2f}<br/>"
            f"📅 Monthly Income<br/>"
            f"₹{financial_summary['monthly_income']:,.2f}<br/>"
            f"📅 Monthly Expense<br/>"
            f"₹{financial_summary['monthly_expense']:,.2f}<br/>"
            f"🎯 Budget<br/>"
            f"₹{financial_summary['budget']:,.2f}<br/>"
            f"💳 Remaining Budget<br/>"
            f"₹{financial_summary['remaining_budget']:,.2f}<br/>"
            f"📊 Budget Used<br/>"
            f"{financial_summary['budget_utilization']}%<br/>"
            f"📈 Total Investment<br/>"
            f"₹{financial_summary['invested']:,.2f}<br/>"
            f"💹 Current Value<br/>"
            f"₹{financial_summary['current_value']:,.2f}<br/>"
            f"💰 Investment Profit<br/>"
            f"₹{financial_summary['investment_profit']:,.2f}"
        )
    else:
        reply = "I don't understand. Try Type help."
    return jsonify({
        "reply": reply
    })

@app.route("/feedback", methods=["GET", "POST"])
@login_required
def feedback():
    if request.method == "POST":
        subject = request.form["subject"]
        message = request.form["message"]
        rating = request.form["rating"]
        fb = Feedback(
            user_id=current_user.id,
            subject=subject,
            message=message,
            rating=rating
        )
        db.session.add(fb)
        db.session.commit()
        msg = Message(
            subject=f"Feedback from {current_user.fullname}",
            sender=current_app.config["MAIL_USERNAME"],
            recipients=["anilvulli45@gmail.com"]
        )
        msg.body = f""" New Feedback Received Name : {current_user.fullname} Email : {current_user.email} Rating : {rating}/5 Subject : {subject} Message : {message} """
        current_app.extensions["mail"].send(msg)
        flash("Feedback sent successfully!", "success")
        return redirect(url_for("dashboard"))
    return render_template("feedback.html")
UPLOAD_FOLDER = "static/images"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    if request.method == "POST":
        current_user.fullname = request.form.get("fullname")
        current_user.gender = request.form.get("gender")
        current_user.occupation = request.form.get("occupation")
        income = request.form.get("monthly_income")
        if income:
            current_user.monthly_income = float(income)
        current_user.city = request.form.get("city")
        current_user.address = request.form.get("address")
        image = request.files.get("profile_image")
        if image and image.filename != "":
            extension = image.filename.rsplit(".", 1)[1].lower()
            filename = f"user_{current_user.id}.{extension}"
            filepath = os.path.join(
                app.config["UPLOAD_FOLDER"],
                filename
            )
            image.save(filepath)
            current_user.profile_image = filename
        db.session.commit()
        flash("Profile updated successfully.", "success")
        return redirect(url_for("profile"))
    return render_template("profile.html")

@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "POST":
        current_password = request.form.get("current_password")
        new_password = request.form.get("new_password")
        confirm_password = request.form.get("confirm_password")
        if not current_password or not new_password or not confirm_password:
            flash("All password fields are required.", "danger")
            return redirect(url_for("change_password"))
        if not check_password_hash(
            current_user.password,
            current_password
        ):
            flash("Current password is incorrect.", "danger")
            return redirect(url_for("change_password"))
        if new_password != confirm_password:
            flash("New passwords do not match.", "danger")
            return redirect(url_for("change_password"))
        if len(new_password) < 8:
            flash(
                "New password must contain at least 8 characters.",
                "danger"
            )
            return redirect(url_for("change_password"))
        if check_password_hash(
            current_user.password,
            new_password
        ):
            flash(
                "New password must be different from your current password.",
                "danger"
            )
            return redirect(url_for("change_password"))
        current_user.password = generate_password_hash(
            new_password
        )
        db.session.commit()
        flash(
            "Password changed successfully!",
            "success"
        )
        return redirect(url_for("profile"))
    return render_template("change_password.html")